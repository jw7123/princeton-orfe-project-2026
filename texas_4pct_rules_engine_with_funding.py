
from __future__ import annotations
import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Dict, List
import datetime as dt
from copy import copy

import openpyxl
from openpyxl.styles import Alignment

TERMINAL_STATUSES = {"Closed", "Withdrawn", "Terminated"}

def yn(value: object) -> Optional[bool]:
    if value is None:
        return None
    s = str(value).strip().upper()
    if s == "Y":
        return True
    if s == "N":
        return False
    return None

def is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""

def as_date(value: object) -> Optional[dt.date]:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return dt.datetime.strptime(v, fmt).date()
            except ValueError:
                pass
    return None

def as_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", "").replace("$", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0

@dataclass
class ProjectResult:
    predicted_status: str
    predicted_path: str
    blocker_code: str
    reason_1: str
    reason_2: str
    reason_3: str
    advance_flag: str

MANUAL_FIELDS = [
    "site_control_valid_flag",
    "resolutions_complete_flag",
    "local_approval_complete_flag",
    "third_party_reports_complete_flag",
    "esa_complete_flag",
    "market_analysis_complete_flag",
    "appraisal_complete_flag",
    "scr_complete_flag",
    "feasibility_report_complete_flag",
    "deficiency_notice_issued_flag",
    "deficiency_cleared_flag",
    "material_deficiency_flag",
    "underwriting_complete_flag",
    "underwriting_pass_flag",
    "financial_feasible_flag",
    "deconcentration_compliant_flag",
    "one_mile_rule_compliant_flag",
    "existing_lura_conflict_flag",
    "streamlined_eligible_flag",
    "board_review_required_flag",
    "chapter12_issue_flag",
    "ready_for_determination_notice_flag",
    "ready_for_board_flag",
]

def evaluate_project(row: Dict[str, object]) -> ProjectResult:
    status_current = str(row.get("application_status_current") or "").strip()

    if status_current == "Closed":
        return ProjectResult("Closed", "Terminal", "", "Closed projects persist.", "", "", "Y")
    if status_current in {"Withdrawn", "Terminated"}:
        return ProjectResult(status_current, "Terminal", "", "Terminal project persists.", "", "", "N")

    if yn(row.get("site_control_valid_flag")) is not True:
        return ProjectResult("No Advancement", "Blocked", "SITE_CONTROL",
                             "Site control missing or invalid.", "", "", "N")

    if yn(row.get("resolutions_complete_flag")) is not True or yn(row.get("local_approval_complete_flag")) is not True:
        return ProjectResult("No Advancement", "Blocked", "APPROVALS",
                             "Required approvals or resolutions incomplete.", "", "", "N")

    if yn(row.get("third_party_reports_complete_flag")) is not True:
        return ProjectResult("No Advancement", "Blocked", "REPORTS",
                             "Required third-party reports incomplete.", "", "", "N")

    if yn(row.get("material_deficiency_flag")) is True or (
        yn(row.get("deficiency_notice_issued_flag")) is True and yn(row.get("deficiency_cleared_flag")) is not True
    ):
        return ProjectResult("No Advancement", "Blocked", "DEFICIENCY",
                             "Unresolved deficiency or material deficiency.", "", "", "N")

    if yn(row.get("underwriting_pass_flag")) is not True or yn(row.get("financial_feasible_flag")) is not True:
        return ProjectResult("No Advancement", "Blocked", "UNDERWRITING",
                             "Underwriting / feasibility not passed.", "", "", "N")

    if (
        yn(row.get("deconcentration_compliant_flag")) is not True
        or yn(row.get("one_mile_rule_compliant_flag")) is not True
        or yn(row.get("existing_lura_conflict_flag")) is True
        or yn(row.get("chapter12_issue_flag")) is True
    ):
        return ProjectResult("No Advancement", "Blocked", "COMPLIANCE",
                             "Deconcentration, one-mile, LURA, or Chapter 12 issue.", "", "", "N")

    if yn(row.get("streamlined_eligible_flag")) is True and yn(row.get("ready_for_determination_notice_flag")) is True and yn(row.get("board_review_required_flag")) is not True:
        return ProjectResult(
            "Determination Notice Issued",
            "Determination Notice",
            "",
            "Streamlined path assumptions satisfied.",
            "No board review required.",
            "",
            "Y",
        )

    if yn(row.get("ready_for_board_flag")) is True or yn(row.get("board_review_required_flag")) is True:
        return ProjectResult(
            "Board",
            "Board",
            "",
            "Board-required or board-ready assumptions satisfied.",
            "",
            "",
            "Y",
        )

    return ProjectResult(
        "Under Review",
        "Under Review",
        "",
        "No hard blocker identified, but project is not yet DN-ready or Board-ready.",
        "",
        "",
        "N",
    )

def load_headers(ws, header_row: int = 3) -> Dict[str, int]:
    return {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}

def load_parameters(ws) -> Dict[str, object]:
    params: Dict[str, object] = {}
    for row_idx in range(3, ws.max_row + 1):
        key = ws.cell(row_idx, 1).value
        if key:
            params[str(key)] = ws.cell(row_idx, 2).value
    return params

def ensure_parameter_rows(ws) -> None:
    existing = {str(ws.cell(r, 1).value): r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value}
    template_row = 3
    new_params = [
        ("allocation_volume_cap_amount", 300000000, "Y", "Planning assumption for next funding cycle capacity; adjust manually."),
        ("allocate_only_dn_or_board_ready_flag", "Y", "Y", "Y = only projects predicted to Board or Determination Notice enter allocation queue."),
        ("allocation_priority_rule", "bond_reservation_date_then_pipeline_entry_then_manual_rank", "Y", "Based on QAP priority concept; tie-break assumptions are modeler-defined."),
        ("exclude_terminal_statuses_from_allocation", "Y", "Y", "Closed / Withdrawn / Terminated do not consume new-cycle capacity."),
        ("carryforward_projects_sort_first", "Y", "Y", "Carryforward projects sort first when all else is equal."),
    ]
    for key, value, use, note in new_params:
        if key in existing:
            row_idx = existing[key]
        else:
            row_idx = ws.max_row + 1
            for c in range(1, 5):
                src = ws.cell(template_row, c)
                dst = ws.cell(row_idx, c)
                dst._style = copy(src._style)
                dst.number_format = src.number_format
        ws.cell(row_idx, 1).value = key
        ws.cell(row_idx, 2).value = value
        ws.cell(row_idx, 3).value = use
        ws.cell(row_idx, 4).value = note
        if isinstance(value, (int, float)):
            ws.cell(row_idx, 2).number_format = '$#,##0_);[Red]($#,##0)'

def write_rules_output(wb, headers: Dict[str, int], projects: List[Dict[str, object]]) -> None:
    out = wb["Rules_Output"]
    header_style = copy(out["A2"]._style)
    row_out = 3
    for project in [p for p in projects if p["sheet_name"] == "Applicant_List"]:
        row = project["row"]
        result = project["result"]
        out.cell(row_out, 1).value = row.get("tdhca_number")
        out.cell(row_out, 2).value = row.get("development_name")
        out.cell(row_out, 3).value = row.get("application_status_current")
        out.cell(row_out, 4).value = result.predicted_status
        out.cell(row_out, 5).value = result.predicted_path
        out.cell(row_out, 6).value = result.advance_flag
        out.cell(row_out, 7).value = result.blocker_code
        out.cell(row_out, 8).value = result.reason_1
        manual_missing = 0
        for field in MANUAL_FIELDS:
            val = row.get(field)
            if val is None or str(val).strip() == "" or str(val).strip().upper() == "N":
                manual_missing += 1
        out.cell(row_out, 9).value = manual_missing
        row_out += 1

    extra_headers = {
        "M2": "allocation_eligible_flag",
        "N2": "simulated_queue_rank",
        "O2": "requested_bond_amount",
        "P2": "simulated_funded_flag",
        "Q2": "simulated_funding_reason",
    }
    for coord, text in extra_headers.items():
        out[coord] = text
        out[coord]._style = copy(header_style)
    out.column_dimensions["M"].width = 18
    out.column_dimensions["N"].width = 18
    out.column_dimensions["O"].width = 18
    out.column_dimensions["P"].width = 18
    out.column_dimensions["Q"].width = 48

    row_out = 3
    for project in [p for p in projects if p["sheet_name"] == "Applicant_List"]:
        alloc = project.get("allocation", {})
        if project["status_current"] in TERMINAL_STATUSES:
            reason = "Terminal / already resolved project; excluded from new allocation cycle."
        elif project["result"].advance_flag != "Y":
            reason = project["result"].reason_1 or "Not eligible under current rules engine."
        else:
            reason = alloc.get("reason", "Advances in process but not modeled as allocation-ready under current assumptions.")
        out.cell(row_out, 13).value = "Y" if project["allocation_eligible"] else "N"
        out.cell(row_out, 14).value = alloc.get("queue_rank")
        out.cell(row_out, 15).value = project["request"] if project["request"] else None
        out.cell(row_out, 16).value = alloc.get("funded", "N")
        out.cell(row_out, 17).value = reason
        row_out += 1

def write_hypothetical_output(wb, project: Dict[str, object]) -> None:
    ws = wb["Hypothetical_Project"]
    headers = [
        "allocation_eligible_flag",
        "allocation_priority_date",
        "simulated_queue_rank",
        "simulated_funded_flag",
        "simulated_remaining_capacity_after",
        "simulated_funding_reason",
    ]
    start_col = 87
    for idx, header in enumerate(headers, start=start_col):
        ws.cell(3, idx).value = header
        ws.cell(3, idx)._style = copy(ws.cell(3, 1)._style)

    alloc = project.get("allocation", {})
    values = [
        "Y" if project["allocation_eligible"] else "N",
        project["priority_date"],
        alloc.get("queue_rank"),
        alloc.get("funded", "N"),
        alloc.get("remaining_after"),
        alloc.get("reason", project["result"].reason_1 if project["result"].advance_flag != "Y" else "Advances in process but not modeled as allocation-ready under current assumptions."),
    ]
    for idx, value in enumerate(values, start=start_col):
        cell = ws.cell(4, idx)
        cell.value = value
        cell._style = copy(ws.cell(4, 1)._style)
    ws.column_dimensions["CI"].width = 18
    ws.column_dimensions["CJ"].width = 14
    ws.column_dimensions["CK"].width = 16
    ws.column_dimensions["CL"].width = 16
    ws.column_dimensions["CM"].width = 22
    ws.column_dimensions["CN"].width = 52

def write_funding_output(wb, projects: List[Dict[str, object]], capacity: float, remaining: float) -> None:
    if "Funding_Output" in wb.sheetnames:
        idx = wb.sheetnames.index("Funding_Output")
        del wb["Funding_Output"]
        ws = wb.create_sheet("Funding_Output", idx)
    else:
        ws = wb.create_sheet("Funding_Output")

    title_style = copy(wb["Applicant_List"]["A1"]._style)
    section_style = copy(wb["Applicant_List"]["A2"]._style)
    header_style = copy(wb["Rules_Output"]["A2"]._style)
    body_style = copy(wb["Rules_Output"]["A3"]._style)

    ws["A1"] = "Funding Output / Queue + Capacity Simulator"
    ws["A1"]._style = title_style
    ws.merge_cells("A1:N1")
    ws["A2"] = "Summary"
    ws["A2"]._style = section_style

    summary_labels = [
        "Capacity assumption",
        "Eligible projects in queue",
        "Projects funded in simulation",
        "Total requested by funded projects",
        "Remaining unallocated capacity",
        "Hypothetical project funded?",
    ]
    for i, label in enumerate(summary_labels, start=3):
        ws.cell(i, 1).value = label
        ws.cell(i, 1)._style = body_style
        ws.cell(i, 2)._style = body_style

    table_start = 11
    headers_out = [
        "queue_rank", "source", "tdhca_number", "development_name", "status_current",
        "status_predicted", "advance_flag", "allocation_eligible_flag", "priority_date",
        "requested_bond_amount", "cumulative_allocated_if_funded", "simulated_funded_flag",
        "remaining_capacity_after", "funding_reason"
    ]
    for c, header in enumerate(headers_out, start=1):
        ws.cell(table_start, c).value = header
        ws.cell(table_start, c)._style = header_style

    ordered = sorted(
        projects,
        key=lambda p: (
            0 if p["source"] == "Hypothetical" else 1,
            p.get("allocation", {}).get("queue_rank", 999999),
            0 if p["allocation_eligible"] else 1,
            str(p["development_name"] or ""),
        ),
    )

    row_out = table_start + 1
    for project in ordered:
        alloc = project.get("allocation", {})
        if project["status_current"] in TERMINAL_STATUSES:
            reason = "Terminal / already resolved project; excluded from new allocation cycle."
        elif project["result"].advance_flag != "Y":
            reason = project["result"].reason_1 or "Not eligible under current rules engine."
        else:
            reason = alloc.get("reason", "Advances in process but not modeled as allocation-ready under current assumptions.")

        values = [
            alloc.get("queue_rank"),
            project["source"],
            project["tdhca_number"],
            project["development_name"],
            project["status_current"],
            project["result"].predicted_status,
            project["result"].advance_flag,
            "Y" if project["allocation_eligible"] else "N",
            project["priority_date"],
            project["request"] if project["request"] else None,
            alloc.get("cumulative"),
            alloc.get("funded", "N"),
            alloc.get("remaining_after"),
            reason,
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(row_out, c).value = value
            ws.cell(row_out, c)._style = body_style
        if project["source"] == "Hypothetical":
            for c in range(1, 15):
                ws.cell(row_out, c).fill = copy(wb["Applicant_List"]["A2"].fill)
                ws.cell(row_out, c).font = copy(wb["Applicant_List"]["A2"].font)
        ws.cell(row_out, 14).alignment = Alignment(wrap_text=True, vertical="top")
        row_out += 1

    funded_count = sum(1 for p in projects if p.get("allocation", {}).get("funded") == "Y")
    funded_total = sum(p["request"] for p in projects if p.get("allocation", {}).get("funded") == "Y")
    hypo = next((p for p in projects if p["source"] == "Hypothetical"), None)

    ws["B3"] = capacity
    ws["B4"] = sum(1 for p in projects if p["allocation_eligible"])
    ws["B5"] = funded_count
    ws["B6"] = funded_total
    ws["B7"] = remaining
    ws["B8"] = hypo.get("allocation", {}).get("funded", "N") if hypo else "N"

    widths = {
        "A": 10, "B": 16, "C": 14, "D": 30, "E": 16, "F": 24, "G": 12,
        "H": 18, "I": 14, "J": 18, "K": 22, "L": 16, "M": 20, "N": 60,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = f"A{table_start + 1}"
    ws.auto_filter.ref = f"A{table_start}:N{row_out - 1}"

def main() -> None:
    parser = argparse.ArgumentParser(description="Run Texas 4% bond rules engine plus queue-based funding layer.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default="Applicant_List")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.workbook)
    applicant_ws = wb[args.sheet]
    headers = load_headers(applicant_ws, 3)

    ensure_parameter_rows(wb["Engine_Parameters"])
    params = load_parameters(wb["Engine_Parameters"])
    capacity = as_number(params.get("allocation_volume_cap_amount", 300000000))

    projects: List[Dict[str, object]] = []
    for sheet_name, row_idx, source in (
        [("Applicant_List", r, "Existing Pipeline") for r in range(4, wb["Applicant_List"].max_row + 1)]
        + [("Hypothetical_Project", 4, "Hypothetical")]
    ):
        ws = wb[sheet_name]
        if sheet_name == "Applicant_List":
            if is_blank(ws.cell(row_idx, headers["tdhca_number"]).value) and is_blank(ws.cell(row_idx, headers["development_name"]).value):
                continue
        else:
            if is_blank(ws.cell(row_idx, headers["development_name"]).value):
                continue

        row = {name: ws.cell(row_idx, col).value for name, col in headers.items()}
        result = evaluate_project(row)
        status_current = str(row.get("application_status_current") or "").strip()
        priority_date = as_date(row.get("bond_reservation_date_current")) or as_date(row.get("pipeline_entry_date")) or as_date(row.get("full_application_submission_date"))
        pipeline_entry = as_date(row.get("pipeline_entry_date")) or as_date(row.get("full_application_submission_date"))
        request = as_number(row.get("bond_reservation_amount_requested")) or as_number(row.get("bond_amount_recommended"))
        allocation_eligible = (
            result.advance_flag == "Y"
            and result.predicted_status in {"Determination Notice Issued", "Board"}
            and status_current not in TERMINAL_STATUSES
        )
        manual_rank = row.get("pipeline_seniority_rank_manual")
        manual_rank = int(manual_rank) if isinstance(manual_rank, (int, float)) and not isinstance(manual_rank, bool) else 999999

        projects.append({
            "sheet_name": sheet_name,
            "row_idx": row_idx,
            "source": source,
            "tdhca_number": row.get("tdhca_number"),
            "development_name": row.get("development_name"),
            "status_current": status_current,
            "priority_date": priority_date,
            "pipeline_entry": pipeline_entry,
            "request": request,
            "carry": yn(row.get("carryforward_from_prior_log_flag")) is True,
            "manual_rank": manual_rank,
            "allocation_eligible": allocation_eligible,
            "row": row,
            "result": result,
        })

    eligible = sorted(
        [p for p in projects if p["allocation_eligible"]],
        key=lambda p: (
            0 if p["carry"] else 1,
            p["priority_date"] or dt.date(9999, 12, 31),
            p["pipeline_entry"] or dt.date(9999, 12, 31),
            p["manual_rank"],
            str(p["tdhca_number"] or ""),
            str(p["development_name"] or ""),
            0 if p["source"] == "Existing Pipeline" else 1,
        ),
    )

    remaining = capacity
    cumulative = 0.0
    for i, project in enumerate(eligible, start=1):
        funded = project["request"] > 0 and project["request"] <= remaining
        if funded:
            cumulative += project["request"]
            remaining -= project["request"]
            reason = "Funded within current simulated capacity."
        elif project["request"] <= 0:
            reason = "No requested bond amount entered."
        else:
            reason = "Eligible, but insufficient remaining capacity in current simulated cycle."
        project["allocation"] = {
            "queue_rank": i,
            "funded": "Y" if funded else "N",
            "remaining_after": remaining,
            "cumulative": cumulative,
            "reason": reason,
        }

    for project in projects:
        project.setdefault("allocation", {})

    write_rules_output(wb, headers, projects)
    write_hypothetical_output(wb, next(p for p in projects if p["source"] == "Hypothetical"))
    write_funding_output(wb, projects, capacity, remaining)

    wb["Read_Me"]["A16"] = "New in this version"
    wb["Read_Me"]["B16"] = "Adds a queue-based funding allocation layer: projects are screened by the existing rules engine, sorted by reservation-date priority, and funded until the manual capacity assumption is exhausted."

    wb.save(args.workbook)

if __name__ == "__main__":
    main()
