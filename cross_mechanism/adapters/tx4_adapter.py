"""
adapters/tx4_adapter.py
=======================

Read the populated `Funding_Output` sheet of the TX 4% workbook
(produced by `texas_4pct_rules_engine_with_funding.py`) and emit a CSV
in the cross-mechanism schema.

This adapter intentionally does NOT re-run the simulator. The logic for
deriving the queue, the priority tiers, and the funding decisions lives
in `texas_4pct_rules_engine_with_funding.py`; this adapter only
translates the simulator's output into the standardized schema so the
harness can compare it to the other mechanisms.

If `Funding_Output` is missing or empty in the supplied workbook, run
the simulator first:

    python texas_4pct_rules_engine_with_funding.py <workbook>

then re-run this adapter against the updated workbook.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from cross_mechanism.schema import CROSS_MECHANISM_COLUMNS, validate_schema


HEADER_ROW = 11   # Funding_Output table starts at row 11 per simulator code


def load_funding_output(workbook_path: Path) -> pd.DataFrame:
    """Read the Funding_Output sheet into a DataFrame."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Funding_Output" not in wb.sheetnames:
        raise ValueError(
            f"Workbook {workbook_path} has no Funding_Output sheet. "
            f"Run texas_4pct_rules_engine_with_funding.py first."
        )
    ws = wb["Funding_Output"]
    headers = [ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        row = {h: ws.cell(r, c + 1).value for c, h in enumerate(headers) if h}
        if not any(v is not None and str(v).strip() != "" for v in row.values()):
            continue
        rows.append(row)
    return pd.DataFrame(rows)


def to_schema(funding_df: pd.DataFrame, cycle_year: int) -> pd.DataFrame:
    """Convert Funding_Output rows into the cross-mechanism schema."""
    out_rows = []
    for _, r in funding_df.iterrows():
        # queue_rank is a within-2025-cohort lower-is-better ordinal
        queue_rank = r.get("queue_rank")
        try:
            score_or_priority = float(queue_rank) if queue_rank is not None else float("nan")
        except (ValueError, TypeError):
            score_or_priority = float("nan")

        # Map the simulator's funding_reason to a blocker code
        funded = str(r.get("simulated_funded_flag") or "N").strip().upper() == "Y"
        status = str(r.get("status_current") or "").strip()
        if funded:
            blocked_by = ""
        elif status in {"Closed", "Withdrawn", "Terminated"}:
            blocked_by = f"terminal:{status.lower()}"
        else:
            blocked_by = "capacity_exhausted"

        request_amount = r.get("requested_bond_amount") or 0.0
        try:
            request_amount = float(request_amount)
        except (ValueError, TypeError):
            request_amount = 0.0

        out_rows.append({
            "mechanism":            "TX_4pct",
            "cycle_year":           cycle_year,
            "app_id":               str(r.get("tdhca_number") or r.get("development_name") or ""),
            "name":                 str(r.get("development_name") or ""),
            "regional_pool":        "",
            "set_aside_flags":      "",
            "score_or_priority":    score_or_priority,
            "request_amount":       request_amount,
            "predicted_award":      request_amount if funded else 0.0,
            "predicted_via":        str(r.get("status_predicted") or ""),
            "predicted_blocked_by": blocked_by,
            "actual_award":         float("nan"),
        })
    return pd.DataFrame(out_rows, columns=CROSS_MECHANISM_COLUMNS)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", type=Path, required=True,
                    help="Texas_4pct_Bond_Simulator_with_Funding_Layer.xlsx")
    ap.add_argument("--cycle-year", type=int, default=2025)
    ap.add_argument("--out", type=Path, required=True,
                    help="Output CSV path.")
    args = ap.parse_args()

    args.out.parent.mkdir(parents=True, exist_ok=True)
    raw = load_funding_output(args.workbook)
    schema_df = to_schema(raw, args.cycle_year)

    # Drop the synthetic "Hypothetical" row injected by the simulator's
    # interactive mode; the cross-mechanism sweep handles synthetics
    # separately via sweep_runner.py.
    schema_df = schema_df[~schema_df["name"].str.contains("Jimbo|Hypothetical|hypo", case=False, na=False, regex=True)]

    validate_schema(schema_df)
    schema_df.to_csv(args.out, index=False)
    n_funded = int((schema_df["predicted_award"] > 0).sum())
    print(f"Wrote {len(schema_df)} rows ({n_funded} funded) to {args.out}; schema OK.")


if __name__ == "__main__":
    main()
